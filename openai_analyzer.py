#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OpenAI API 公告分析器
針對 sbj_pu11 表中 RULC IN (1,11) 的公告進行結構化分析
生成摘要、when、how_much、who_what 四種 Excel 格式解析
"""

import openai
import pandas as pd
import json
import logging
import argparse
from datetime import datetime
from create_mysql_db import MySQLHandler
import configparser
import tiktoken
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

class OpenAIAnalyzer:
    def __init__(self, config_file='config.ini', test_mode=False):
        self.config_file = config_file
        self.test_mode = test_mode
        self.load_config()
        self.setup_openai()
        
        # 限速處理相關變數
        self.rate_limit_count = 0  # 累計限速次數
        self.consecutive_success = 0  # 連續成功次數
        self.max_rate_limit_attempts = 4  # 最大重試次數（第4次限速時停止程式）
        self.base_wait_time = 60  # 基礎等待時間（秒）
        self.reset_threshold = 5  # 連續成功5次後重置限速計數
        
    def load_config(self):
        """載入設定檔"""
        config = configparser.ConfigParser()
        config.read(self.config_file, encoding='utf-8')
        self.openai_api_key = config.get('openai', 'api_key')
        
    def setup_openai(self):
        """設定 OpenAI API"""
        self.client = openai.OpenAI(api_key=self.openai_api_key)
        # 初始化 tiktoken 編碼器
        self.encoding = tiktoken.encoding_for_model("gpt-3.5-turbo")
    
    def estimate_tokens(self, text):
        """估算文本的 token 數量"""
        try:
            return len(self.encoding.encode(text))
        except Exception as e:
            # 如果編碼失敗，使用簡單的字元數估算（大約 4 字元 = 1 token）
            logging.warning(f"Token 計算失敗，使用字元數估算: {e}")
            return len(text) // 4
    
    def select_model(self, content, analysis_type="general"):
        """根據內容長度動態選擇最適合的模型
        
        Args:
            content: 要分析的內容
            analysis_type: 分析類型，用於決定輸出長度
            
        Returns:
            tuple: (model_name, max_tokens)
        """
        # 計算輸入 token 數
        input_tokens = self.estimate_tokens(content)
        
        # 根據分析類型設定預期輸出 token 數
        output_token_map = {
            "summary": 300,      # 摘要較短
            "when": 600,         # 時間資訊中等
            "how_much": 800,     # 數量金額較長
            "who_what": 800,     # 人物關係較長
            "general": 600       # 一般情況
        }
        
        expected_output_tokens = output_token_map.get(analysis_type, 600)
        
        # 總需求 token 數（輸入 + 輸出 + 系統訊息緩衝）
        total_tokens_needed = input_tokens + expected_output_tokens + 200
        
        # 根據 token 需求選擇模型（從 gpt-4-turbo 開始使用）
        if total_tokens_needed <= 120000:  # 128k 的緩衝
            model = "gpt-4-turbo"
            max_context = 128000
            logging.info(f"選擇 gpt-4-turbo (預估需求: {total_tokens_needed} tokens)")
        else:
            model = "gpt-4o-mini"  # 使用 OpenAI 實際可用的大容量模型
            max_context = 128000   # gpt-4o-mini 實際上下文長度
            logging.info(f"選擇 gpt-4o-mini (預估需求: {total_tokens_needed} tokens)")
            
        return model, min(expected_output_tokens, 4000)  # 限制最大輸出長度
    
    def handle_rate_limit(self):
        """處理限速情況的智能等待機制"""
        self.rate_limit_count += 1
        self.consecutive_success = 0  # 重置連續成功計數
        
        # 檢查是否達到最大重試次數
        if self.rate_limit_count >= self.max_rate_limit_attempts:
            error_msg = f"已達到最大限速重試次數 ({self.max_rate_limit_attempts} 次)，程式將停止執行。"
            logging.error(error_msg)
            raise Exception(error_msg)
        
        # 計算等待時間（指數退避：第1次60秒，第2次120秒，第3次240秒...）
        wait_time = self.base_wait_time * (2 ** (self.rate_limit_count - 1))
        wait_minutes = wait_time / 60
        
        logging.warning(f"遇到限速 (第 {self.rate_limit_count} 次)，將等待 {wait_minutes:.1f} 分鐘後重試...")
        logging.info(f"預計恢復時間: {(datetime.now() + pd.Timedelta(seconds=wait_time)).strftime('%H:%M:%S')}")
        
        # 分段等待，每30秒顯示一次剩餘時間
        remaining_time = wait_time
        while remaining_time > 0:
            sleep_duration = min(30, remaining_time)
            time.sleep(sleep_duration)
            remaining_time -= sleep_duration
            
            if remaining_time > 0:
                remaining_minutes = remaining_time / 60
                logging.info(f"還需等待 {remaining_minutes:.1f} 分鐘...")
        
        logging.info("等待結束，準備重新嘗試...")
    
    def record_success(self):
        """記錄成功執行，用於重置限速計數"""
        self.consecutive_success += 1
        
        # 如果連續成功達到閾值，重置限速計數
        if self.consecutive_success >= self.reset_threshold:
            if self.rate_limit_count > 0:
                logging.info(f"連續成功 {self.reset_threshold} 次，重置限速計數 (之前: {self.rate_limit_count} 次)")
                self.rate_limit_count = 0
            self.consecutive_success = 0
    
    def call_openai_with_retry(self, model, messages, max_tokens, temperature, analysis_type="未知"):
        """帶有限速重試機制的 OpenAI API 呼叫"""
        max_attempts = 3  # 每次分析的最大重試次數
        
        for attempt in range(max_attempts):
            try:
                response = self.client.chat.completions.create(
                    model=model,
                    messages=messages,
                    max_tokens=max_tokens,
                    temperature=temperature
                )
                
                # 成功執行，記錄成功
                self.record_success()
                return response
                
            except openai.RateLimitError as e:
                logging.error(f"{analysis_type} 分析遇到限速錯誤: {e}")
                
                if attempt < max_attempts - 1:
                    # 還有重試機會，處理限速
                    self.handle_rate_limit()
                    logging.info(f"重試 {analysis_type} 分析 (第 {attempt + 2} 次嘗試)")
                else:
                    # 已達到最大重試次數
                    logging.error(f"{analysis_type} 分析達到最大重試次數，跳過此次分析")
                    raise e
                    
            except Exception as e:
                # 其他錯誤直接拋出
                logging.error(f"{analysis_type} 分析發生錯誤: {e}")
                raise e
        
        # 理論上不會到達這裡
        raise Exception(f"{analysis_type} 分析失敗")
    
    def log_openai_conversation(self, analysis_type, prompt, response, output_dir, file_prefix, model_used=None):
        """記錄 OpenAI 對話內容到檔案"""
        if not self.test_mode:
            return
            
        log_content = f"""=== OpenAI {analysis_type.upper()} 分析對話記錄 ===

【使用模型】
{model_used or '未指定'}

【發送給 OpenAI 的 Prompt】
{prompt}

【OpenAI 的回應】
{response}

【分析時間】
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

{'='*60}
"""
        
        with open(f"{output_dir}/{file_prefix}_{analysis_type}_openai_log.txt", 'w', encoding='utf-8') as f:
            f.write(log_content)
        
    def get_cl1_announcements(self, mysql_handler, limit=None, rulc_values=None):
        """從資料庫取得指定 RULC 的公告，並對應 tej_pu11_1 表的 txt 內容
        
        Args:
            mysql_handler: MySQL 連接處理器
            limit: 限制筆數
            rulc_values: RULC 值列表，預設為 [1, 11]
        """
        if not mysql_handler.connection:
            logging.error("MySQL 未連接")
            return None
        
        # 設定預設的 RULC 值
        if rulc_values is None:
            rulc_values = [1, 11]
        
        # 確保 rulc_values 是列表格式
        if not isinstance(rulc_values, list):
            rulc_values = [rulc_values]
            
        try:
            cursor = mysql_handler.connection.cursor(dictionary=True)
            
            # 構建 RULC 的 IN 條件
            rulc_placeholders = ','.join(['%s'] * len(rulc_values))
            
            # 先查詢 sbj_pu11 中指定 rulc 且尚未處理的公告，包含所有需要的欄位
            sql = f"""
            SELECT id, ban, code, name, d_reals, hr_reals, od, cl, rulc
            FROM sbj_pu11 
            WHERE rulc IN ({rulc_placeholders})
            AND openai_processed = 0 
            ORDER BY d_reals DESC, hr_reals DESC
            """
            if limit:
                sql += f" LIMIT {limit}"
                
            cursor.execute(sql, rulc_values)
            announcements = cursor.fetchall()
            
            # 對每個公告查詢對應的 tej_pu11_1 內容並合併
            results = []
            for announcement in announcements:
                content_sql = """
                SELECT txt 
                FROM tej_pu11_1 
                WHERE code = %s 
                AND d_reals = %s 
                AND hr_reals = %s 
                AND od = %s 
                AND txt IS NOT NULL
                ORDER BY id
                """
                cursor.execute(content_sql, (
                    announcement['code'], 
                    announcement['d_reals'], 
                    announcement['hr_reals'], 
                    announcement['od']
                ))
                content_rows = cursor.fetchall()
                
                if content_rows:
                    # 將所有 txt 內容合併
                    combined_content = '\n'.join([row['txt'] for row in content_rows if row['txt']])
                    announcement['content'] = combined_content
                    results.append(announcement)
            
            logging.info(f"找到 {len(results)} 筆 RULC IN ({','.join(map(str, rulc_values))}) 且尚未處理的公告有對應的 tej_pu11_1 內容")
            return results
            
        except Exception as e:
            logging.error(f"查詢 RULC IN ({','.join(map(str, rulc_values))}) 公告失敗: {e}")
            return None

    def analyze_summary(self, content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir=None, file_prefix=None):
        """生成摘要分析"""
        # 動態選擇模型
        model, max_tokens = self.select_model(content, "summary")
        
        prompt = f"""
【公告基本資訊】
公告ID: {ann_id}
統一編號(BAN): {ban}
公司代碼: {code}
公司名稱: {name}
發言日期(D_REALS): {d_reals}
發言時間(HR_REALS): {hr_reals}
公告序號(OD): {od}
法規條款(RULC): {rulc}

【公告內容】
{content}
"""
        
        try:
            # 設定 system message for 摘要分析
            system_message = {
                "role": "system", 
                "content": """你是一位專業的證券分析師，專門分析台灣證交所公告。你的任務是從複雜的公告內容中提取關鍵資訊，生成精確、完整的摘要。

**核心任務：**
請分析證交所公告內容，提供詳細而精確的摘要。請採用「全面理解」方式，涵蓋所有重要資訊要素。

**摘要內容必須包含以下關鍵要素：**

1. **交易性質和類型**：明確說明是何種類型的交易（如：股權投資、資產收購、合作協議等）

2. **主要當事人**：
   - 交易主體（買方/投資方）
   - 交易對象（賣方/被投資方）
   - 投資標的（標的公司/資產）

3. **核心數據**：
   - 主要交易金額
   - 重要持股比例變化
   - 關鍵財務比率

4. **交易目的和影響**：
   - 策略目的和預期效益
   - 對公司營運的影響
   - 風險因素說明

5. **執行時程**：
   - 重要時間節點
   - 預計完成時間

**摘要品質要求：**
- 長度控制在150-250字
- 用詞精確，避免模糊表達
- 包含具體數字和比例
- 突出最重要的商業意義
- 使用專業但易懂的語言

**摘要結構建議：**
第一句：公司基本資訊和交易性質
第二句：交易標的和主要條件
第三句：核心數據（金額、比例等）
第四句：交易目的和預期影響
第五句：執行時程和注意事項

**執行要求：**
1. 準確識別交易性質和當事人
2. 提取核心數據和關鍵比例  
3. 說明交易目的和商業影響
4. 使用專業但易懂的語言
5. 確保摘要內容具有實用價值

請以繁體中文回應，格式為純文字摘要，不要分段。
確保摘要內容完整、準確、具有實用價值。"""
            }
            
            messages = [
                system_message,
                {"role": "user", "content": prompt}
            ]
            response = self.call_openai_with_retry(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=0.4,
                analysis_type="摘要"
            )
            result = response.choices[0].message.content.strip()
            
            # 記錄對話內容（如果是測試模式）
            if self.test_mode and output_dir and file_prefix:
                self.log_openai_conversation("summary", prompt, result, output_dir, file_prefix, model)
                
            return result
        except Exception as e:
            logging.error(f"摘要分析失敗: {e}")
            return "分析失敗"

    def analyze_when(self, content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir=None, file_prefix=None):
        """分析時間相關資訊"""
        # 動態選擇模型
        model, max_tokens = self.select_model(content, "when")
        
        prompt = f"""
【公告基本資訊】
公告ID: {ann_id}
統一編號(BAN): {ban}
公司代碼: {code}
公司名稱: {name}
發言日期(D_REALS): {d_reals}
發言時間(HR_REALS): {hr_reals}
公告序號(OD): {od}
法規條款(RULC): {rulc}

【公告內容】
{content}
"""
        
        try:
            # 設定 system message for 時間分析
            system_message = {
                "role": "system", 
                "content": """你是一位專業的資料提取專家，專門從證交所公告中提取時間資訊。你必須採用「地毯式搜索」方式，確保無任何時間資訊遺漏。

**核心任務：**
請分析證交所公告內容，提取所有時間相關資訊，以CSV格式輸出。

**全面搜索以下所有類型的時間資訊：**

**基本時間資訊：**
- 發言日期、發言時間
- 事實發生日、事實發生時間
- 公告日期、公告時間
- 申報日期、申報時間

**交易相關時間：**
- 交易日期、交易時間
- 成交日期、成交時間
- 簽約日期、簽約時間
- 合約日期、合約期間
- 交易期間（起始、結束）
- 交割日期、交割期間
- 付款日期、收款日期

**會議時間：**
- 董事會決議日期、董事會會議時間
- 股東會日期、股東會時間
- 法說會日期、法說會時間
- 其他會議日期和時間

**期限和截止時間：**
- 申請期限、申報期限
- 繳款期限、認購期限
- 執行期限、完成期限
- 到期日、截止日
- 生效日期、失效日期

**財務報表相關時間：**
- 財報基準日、財報截止日
- 會計年度、會計期間
- 查核日期、核准日期
- 發布日期、公布日期

**其他重要時間：**
- 預計完成日期、預計執行時間
- 開始營運日期、營運期間
- 投資期間、持有期間
- 任何其他明確的日期或時間

**輸出格式要求：**
請參考以下格式輸出CSV，每個欄位都必須用雙引號包圍：
"項目說明","日期時間"
"發言日期","113/07/09"
"發言時間","21:34:40"
"事實發生日","113/07/09"
"董事會決議日","113/07/01"
"簽約日期","2024年7月5日"
"交易期間（起始）","113/07/08"
"交易期間（結束）","113/07/09"
"交割預定日","113/07/15"
"付款期限","合約簽署後30日內"
"預計完成日","113年第3季"
"財報基準日","113/06/30"
"股東會召開日","113/08/15 上午9時"

**重要注意事項：**
1. **所有欄位都必須用雙引號包圍，包括表頭**
2. 完全保持原文格式，不進行任何轉換
3. 民國年保持民國年格式（如：113/07/09、民國113年7月9日）
4. 西元年保持西元年格式（如：2024/07/09、2024年7月9日）
5. 保留原文的時間表達方式（如：上午、下午、AM、PM）
6. 保留完整的時間描述（包括年、月、日、時、分、秒）
7. 包含相對時間描述（如：簽約後30天、會計年度結束前等）
8. 包含時間範圍和期間（如：2024年1月至3月、第一季度等）
9. 包含重複提及的相同時間（每次提及都要列出）
10. 注意隱含的時間資訊（如：「昨日」、「本月」、「去年」等）

**執行原則：**
- 仔細檢查每個段落，確保所有日期都被提取
- 寧可多提取也不可遺漏任何時間資訊
- 完全維持原文的日期時間格式，不要進行任何轉換或統一化處理
- **確保每個欄位都用雙引號包圍，避免逗號造成欄位錯亂**

請只輸出CSV格式（所有欄位用雙引號包圍），不要其他說明文字。如果某個時間資訊不存在，請省略該行。"""
            }
            
            messages = [
                system_message,
                {"role": "user", "content": prompt}
            ]
            response = self.call_openai_with_retry(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=0.4,
                analysis_type="時間"
            )
            result = response.choices[0].message.content.strip()
            
            # 記錄對話內容（如果是測試模式）
            if self.test_mode and output_dir and file_prefix:
                self.log_openai_conversation("when", prompt, result, output_dir, file_prefix, model)
                
            return result
        except Exception as e:
            logging.error(f"when 分析失敗: {e}")
            return "項目說明,日期時間\n分析失敗,N/A"

    def analyze_how_much(self, content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir=None, file_prefix=None):
        """分析數量金額相關資訊"""
        # 動態選擇模型
        model, max_tokens = self.select_model(content, "how_much")
        
        prompt = f"""
【公告基本資訊】
公告ID: {ann_id}
統一編號(BAN): {ban}
公司代碼: {code}
公司名稱: {name}
發言日期(D_REALS): {d_reals}
發言時間(HR_REALS): {hr_reals}
公告序號(OD): {od}
法規條款(RULC): {rulc}

【公告內容】
{content}
"""
        
        try:
            # 設定 system message for 數量金額分析
            system_message = {
                "role": "system", 
                "content": """你是專業的財務數據提取專家，從證交所公告中提取所有數量、金額、比率資訊。採用「無遺漏」原則。

**任務：**
分析公告內容，將所有金額、數量、比率做成表格，重複的也要列出。

**搜索範圍：**
- 投資/交易金額、累積投資金額、預估價款
- 持股比例（本次、累積、間接、表決權、預計、交割後等）
- **特別注意：累積持有資訊中的持股比率必須全部提取**
- 財務比例（占總資產、股東權益、母公司權益比例等）
- 營運資金、交易數量、合計金額、任何數值和百分比

**輸出格式：**
CSV欄位：所有欄位都必須用雙引號包圍
"類別","項目名稱","標的物","數值","單位","幣別","備註"

**關鍵要求：**
- 類別：金額、數量、比率（只使用這三種分類）
- 數值：純數字，**絕對禁止千分位逗號**，過濾「約」字，**保留正負號**
- 數值和單位依原樣，不換算（6500萬→6500,萬元 而非 65000000,元）
- 幣別依原文，百分比留空
- 備註：投資關係、交易性質、約字說明、重複原因
- 項目名稱含階段說明（預計、交割後、再次提及等）
- 標的物具體明確（如「TCC Dutch 對 TCAH 股份」）
- **重要：每個標的物的持股比率都要列出，即使數值相同也要分別記錄**
- **重要：仔細檢查「累積持有」、「迄目前為止」等段落中的持股比率**
- **重要：負數保留負號（如：-13797385），不要過濾負號**
- **重要：數值欄位絕對不可使用千分位逗號（如：123456 不是 123,456），避免CSV格式錯亂**
- **重要：所有欄位都必須用雙引號包圍，避免逗號造成欄位錯亂**

**範例：**
"類別","項目名稱","標的物","數值","單位","幣別","備註"
"金額","投資總金額","Cimpor Global Holdings B.V.","6500","萬元","歐元","原文為約6500萬元，已過濾約字"
"比率","持股比率","TCC Dutch 對 TCAH 股份","60","%","","TCC Dutch 持有 TCAH 股份比例"
"比率","持股比率","TCC Oyak Amsterdam Holdings B.V.","60","%","","TCC Dutch 持有 TCAH 股份比例"
"比率","持股比率","Cimpor Portugal Holdings SGPS S.A.","100","%","","TCCE 持有 Cimpor 股份比例"
"數量","累積持有股數","台泥對台泥儲能","600600","仟股","","台泥持有台泥儲能累積股數"
"金額","營運資金","公司整體","-13797385","仟元","新台幣","負數表示營運資金為負值"

請只輸出CSV格式（所有欄位用雙引號包圍），不要其他說明文字。"""
            }
            
            messages = [
                system_message,
                {"role": "user", "content": prompt}
            ]
            response = self.call_openai_with_retry(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=0.4,
                analysis_type="數量金額"
            )
            result = response.choices[0].message.content.strip()
            
            # 記錄對話內容（如果是測試模式）
            if self.test_mode and output_dir and file_prefix:
                self.log_openai_conversation("how_much", prompt, result, output_dir, file_prefix, model)
                
            return result
        except Exception as e:
            logging.error(f"how_much 分析失敗: {e}")
            return "類別,項目名稱,標的物,數值,單位,幣別,備註\n分析失敗,N/A,N/A,N/A,N/A,N/A,N/A"

    def analyze_who_what(self, content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir=None, file_prefix=None):
        """分析人物關係相關資訊"""
        # 動態選擇模型
        model, max_tokens = self.select_model(content, "who_what")
        
        prompt = f"""
【公告基本資訊】
公告ID: {ann_id}
統一編號(BAN): {ban}
公司代碼: {code}
公司名稱: {name}
發言日期(D_REALS): {d_reals}
發言時間(HR_REALS): {hr_reals}
公告序號(OD): {od}
法規條款(RULC): {rulc}

【公告內容】
{content}
"""
        
        try:
            # 設定 system message for 人物關係分析
            system_message = {
                "role": "system", 
                "content": """你是專業的交易分析師，從證交所公告中提取交易核心要素，以CSV格式輸出。

**任務：提取交易關鍵資訊**
CSV欄位：所有欄位都必須用雙引號包圍
"標的物","買方","賣方","交易人雙方關係","交易金額"

**提取要求：**
- **標的物**：具體說明股權比例、資產名稱、數量單位
- **買方/賣方**：使用完整公司全稱（含"股份有限公司"等，禁用縮寫）
- **雙方關係**：關係人/非關係人交易、母子公司、關聯企業等
- **交易金額**：含幣別的完整金額表達，**金額數字絕對禁止千分位逗號**

**輸出格式範例：**
"標的物","買方","賣方","交易人雙方關係","交易金額"
"Cimpor Global Holdings B.V.普通股100%股權","台灣水泥股份有限公司","OYAK Capital Investments B.V.","非關係人交易","約6500萬歐元"
"A公司普通股50%股權","B股份有限公司","C股份有限公司","關聯企業","新台幣1000萬元"

**執行原則：**
1. 公司名稱使用完整全稱，不可縮寫
2. 多筆交易分行列出
3. 資訊不明確時填入"資訊不明"
4. 保持原文準確性
5. 必須包含 CSV 表頭行
6. **重要：交易金額中的數字絕對不可使用千分位逗號（如：1000000 不是 1,000,000），避免CSV格式錯亂**
7. **重要：所有欄位都必須用雙引號包圍，避免逗號造成欄位錯亂**

請輸出完整的CSV格式（包含表頭，所有欄位用雙引號包圍）。"""
            }
            
            messages = [
                system_message,
                {"role": "user", "content": prompt}
            ]
            response = self.call_openai_with_retry(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=0.4,
                analysis_type="人物關係"
            )
            result = response.choices[0].message.content.strip()
            
            # 記錄對話內容（如果是測試模式）
            if self.test_mode and output_dir and file_prefix:
                self.log_openai_conversation("who_what", prompt, result, output_dir, file_prefix, model)
                
            return result
        except Exception as e:
            logging.error(f"who_what 分析失敗: {e}")
            return "標的物,買方,賣方,交易人雙方關係,交易金額\n分析失敗,N/A,N/A,N/A,N/A"

    def write_to_excel(self, file_path, data_rows, headers, sheet_name="Sheet1"):
        """
        將資料寫入 Excel 檔案
        
        Args:
            file_path (str): Excel 檔案路徑
            data_rows (list): 資料行列表，每行是一個列表
            headers (list): 表頭列表
            sheet_name (str): 工作表名稱
        """
        import os
        try:
            # 檢查檔案是否存在
            if os.path.exists(file_path):
                # 檔案存在，加載現有的工作簿
                workbook = openpyxl.load_workbook(file_path)
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.create_sheet(sheet_name)
            else:
                # 檔案不存在，創建新的工作簿
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = sheet_name
                
                # 寫入表頭
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    # 設定表頭樣式
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # 找到下一個可用的行
            next_row = worksheet.max_row + 1
            
            # 寫入資料行
            for data_row in data_rows:
                for col, value in enumerate(data_row, 1):
                    worksheet.cell(row=next_row, column=col, value=value)
                next_row += 1
            
            # 自動調整欄寬
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大寬度限制為50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存檔案
            workbook.save(file_path)
            logging.info(f"已將 {len(data_rows)} 行資料寫入 Excel 檔案: {file_path}")
            
        except Exception as e:
            logging.error(f"寫入 Excel 檔案失敗 {file_path}: {e}")
    
    def csv_to_excel_data(self, csv_text, ann_id, ban, code, name, d_reals, hr_reals, od, rulc):
        """
        將 CSV 文本轉換為 Excel 資料行
        支援雙引號包圍的 CSV 格式
        """
        import csv
        import io
        
        try:
            lines = csv_text.strip().split('\n')
            data_rows = []
            
            # 跳過表頭，從第二行開始處理
            for line in lines[1:]:
                if line.strip():
                    # 使用 Python 的 csv 模組來正確解析帶引號的 CSV
                    try:
                        # 創建一個 StringIO 物件來模擬檔案
                        csv_reader = csv.reader(io.StringIO(line))
                        csv_fields = next(csv_reader)
                        
                        # 構建完整的資料行（加入公告基本資訊）
                        full_row = [ann_id, ban, code, name, d_reals, hr_reals, od, rulc] + csv_fields
                        data_rows.append(full_row)
                        
                    except Exception as parse_error:
                        logging.warning(f"CSV 行解析失敗: {line}, 錯誤: {parse_error}")
                        # 如果標準 CSV 解析失敗，回退到簡單分割
                        fallback_fields = [field.strip(' "') for field in line.split(',')]
                        full_row = [ann_id, ban, code, name, d_reals, hr_reals, od, rulc] + fallback_fields
                        data_rows.append(full_row)
            
            return data_rows
            
        except Exception as e:
            logging.error(f"CSV 轉換失敗: {e}")
            return []

    def process_announcements(self, announcements, mysql_handler, output_dir="./analysis_output", analysis_types=None):
        """處理公告列表並生成分析結果 - 輸出為 Excel 格式"""
        import os
        os.makedirs(output_dir, exist_ok=True)
        
        # 預設執行所有分析類型
        if analysis_types is None:
            analysis_types = ['summary', 'when', 'how_much', 'who_what']
        
        # 合併檔案路徑 - 改為 Excel 格式
        when_file_path = f"{output_dir}/all_when_analysis.xlsx"
        how_much_file_path = f"{output_dir}/all_how_much_analysis.xlsx"
        who_what_file_path = f"{output_dir}/all_who_what_analysis.xlsx"
        summary_file_path = f"{output_dir}/all_summary_analysis.txt"
        
        # 定義 Excel 表頭
        when_headers = ["公告ID", "BAN", "公司代碼", "公司名稱", "D_REALS", "HR_REALS", "OD", "RULC", "項目說明", "日期時間"]
        how_much_headers = ["公告ID", "BAN", "公司代碼", "公司名稱", "D_REALS", "HR_REALS", "OD", "RULC", "類別", "項目名稱", "標的物", "數值", "單位", "幣別", "備註"]
        who_what_headers = ["公告ID", "BAN", "公司代碼", "公司名稱", "D_REALS", "HR_REALS", "OD", "RULC", "標的物", "買方", "賣方", "交易人雙方關係", "交易金額"]
        
        # 摘要檔案仍然使用文字格式
        summary_file = open(summary_file_path, 'a', encoding='utf-8') if 'summary' in analysis_types else None
        
        try:
            for i, announcement in enumerate(announcements):
                try:
                    ann_id = announcement['id']
                    ban = announcement['ban']
                    code = announcement['code']
                    name = announcement['name']
                    d_reals = announcement['d_reals']
                    hr_reals = announcement['hr_reals']
                    od = announcement['od']
                    rulc = announcement['rulc']
                    content = announcement['content']  # 使用從 tej_pu11_1 表取得的 txt 內容
                    
                    logging.info(f"正在分析公告 {i+1}/{len(announcements)}: {code} {name} (ID: {ann_id})")
                    start_time = datetime.now()
                    
                    # 建立檔案名稱前綴
                    file_prefix = f"{ann_id}_{code}_{d_reals}"
                    
                    # 檢查內容長度並記錄
                    content_length = len(content)
                    estimated_tokens = self.estimate_tokens(content)
                    logging.info(f"公告內容長度: {content_length} 字元")
                    logging.info(f"預估 token 數: {estimated_tokens}")
                    
                    if content_length > 6000:
                        logging.warning(f"公告內容較長，可能會超過 AI 模型限制 (長度: {content_length})")
                    if estimated_tokens > 15000:
                        logging.warning(f"預估 token 數較高: {estimated_tokens}，將使用高容量模型")
                    
                    # 先儲存送給 OpenAI 的原始內容（保留個別檔案）
                    with open(f"{output_dir}/{file_prefix}_original_content.txt", 'w', encoding='utf-8') as f:
                        f.write("=== 送給 OpenAI 的原始內容 ===\n\n")
                        f.write(f"公告ID: {ann_id}\n")
                        f.write(f"BAN: {ban}\n")
                        f.write(f"公司代碼: {code}\n") 
                        f.write(f"公司名稱: {name}\n")
                        f.write(f"D_REALS: {d_reals}\n")
                        f.write(f"HR_REALS: {hr_reals}\n")
                        f.write(f"OD: {od}\n")
                        f.write(f"RULC: {rulc}\n")
                        f.write("=" * 50 + "\n\n")
                        f.write(content)
                        f.write("\n\n" + "=" * 50)
                    
                    # 生成各類分析（傳遞測試模式參數）
                    analysis_count = len(analysis_types)
                    logging.info(f"開始進行 {analysis_count} 種分析: {', '.join(analysis_types)}")
                    
                    # 摘要分析
                    if 'summary' in analysis_types:
                        summary_start = datetime.now()
                        summary = self.analyze_summary(content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir, file_prefix)
                        summary_time = datetime.now() - summary_start
                        logging.info(f"摘要分析完成 ({summary_time.total_seconds():.2f}秒)")
                    
                    # 時間分析
                    if 'when' in analysis_types:
                        when_start = datetime.now()
                        when_csv = self.analyze_when(content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir, file_prefix)
                        when_time = datetime.now() - when_start
                        logging.info(f"時間分析完成 ({when_time.total_seconds():.2f}秒)")
                    
                    # 數量金額分析
                    if 'how_much' in analysis_types:
                        how_much_start = datetime.now()
                        how_much_csv = self.analyze_how_much(content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir, file_prefix)
                        how_much_time = datetime.now() - how_much_start
                        logging.info(f"數量金額分析完成 ({how_much_time.total_seconds():.2f}秒)")
                    
                    # 人物關係分析
                    if 'who_what' in analysis_types:
                        who_what_start = datetime.now()
                        who_what_csv = self.analyze_who_what(content, ann_id, ban, code, name, d_reals, hr_reals, od, rulc, output_dir, file_prefix)
                        who_what_time = datetime.now() - who_what_start
                        logging.info(f"人物關係分析完成 ({who_what_time.total_seconds():.2f}秒)")
                    
                    # 寫入摘要到文字檔案
                    if 'summary' in analysis_types and summary_file and summary:
                        summary_file.write(f"=== 公告 {ann_id} - {ban} - {code} {name} ({d_reals}) ===\n")
                        summary_file.write(f"BAN: {ban}, D_REALS: {d_reals}, HR_REALS: {hr_reals}, OD: {od}, RULC: {rulc}\n")
                        summary_file.write(f"{summary}\n\n")
                        summary_file.flush()
                    
                    # 寫入時間分析到 Excel
                    if 'when' in analysis_types and when_csv:
                        when_data = self.csv_to_excel_data(when_csv, ann_id, ban, code, name, d_reals, hr_reals, od, rulc)
                        if when_data:
                            self.write_to_excel(when_file_path, when_data, when_headers, "時間分析")
                    
                    # 寫入數量金額分析到 Excel
                    if 'how_much' in analysis_types and how_much_csv:
                        how_much_data = self.csv_to_excel_data(how_much_csv, ann_id, ban, code, name, d_reals, hr_reals, od, rulc)
                        if how_much_data:
                            self.write_to_excel(how_much_file_path, how_much_data, how_much_headers, "數量金額分析")
                    
                    # 寫入人物關係分析到 Excel
                    if 'who_what' in analysis_types and who_what_csv:
                        who_what_data = self.csv_to_excel_data(who_what_csv, ann_id, ban, code, name, d_reals, hr_reals, od, rulc)
                        if who_what_data:
                            self.write_to_excel(who_what_file_path, who_what_data, who_what_headers, "交易分析")
                    
                    # 更新資料庫中的處理狀態
                    if mysql_handler.update_openai_processed_status(ann_id, True):
                        processing_time = datetime.now() - start_time
                        logging.info(f"公告 {ann_id} 分析完成並已標記為已處理 (總耗時: {processing_time.total_seconds():.2f}秒)")
                    else:
                        logging.warning(f"公告 {ann_id} 分析完成但狀態更新失敗")
                    
                except Exception as e:
                    logging.error(f"處理公告 {announcement.get('id', 'unknown')} 失敗: {e}")
                    logging.exception("詳細錯誤資訊:")
                    continue
        
        finally:
            # 關閉摘要檔案
            if summary_file:
                summary_file.close()
                
        logging.info(f"所有分析完成！")
        if 'when' in analysis_types:
            logging.info(f"時間分析結果: {when_file_path}")
        if 'how_much' in analysis_types:
            logging.info(f"數量金額分析結果: {how_much_file_path}")
        if 'who_what' in analysis_types:
            logging.info(f"交易分析結果: {who_what_file_path}")
        if 'summary' in analysis_types:
            logging.info(f"摘要分析結果: {summary_file_path}")

def main():
    parser = argparse.ArgumentParser(description="OpenAI 公告分析器")
    parser.add_argument('--config', default='config.ini', help='設定檔路徑')
    parser.add_argument('--limit', type=int, default=None, help='處理公告數量限制，預設為 None (處理全部公告)')
    parser.add_argument('--output-dir', default='./analysis_output', help='輸出目錄')
    parser.add_argument('--test-mode', action='store_true', help='測試模式：記錄所有與OpenAI的對話內容')
    parser.add_argument('--log-file', default=None, help='日誌檔案路徑（預設：自動生成時間戳記檔名）')
    parser.add_argument('--analysis-types', nargs='+', 
                        choices=['summary', 'when', 'how_much', 'who_what'],
                        default=['summary', 'when', 'how_much', 'who_what'],
                        help='指定要執行的分析類型，可選：summary, when, how_much, who_what')
    parser.add_argument('--rulc', nargs='+', type=int, default=[1, 11],
                        help='指定要查詢的 RULC 值，可輸入多個值，預設為 1 11')
    args = parser.parse_args()
    
    # 設定日誌檔案路徑
    if args.log_file is None:
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.log_file = f"./logs/openai_analyzer_{timestamp}.log"
    
    # 建立日誌目錄
    import os
    log_dir = os.path.dirname(args.log_file) or './logs'
    os.makedirs(log_dir, exist_ok=True)
    
    # 設定日誌記錄器
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s: %(message)s",
        handlers=[
            logging.FileHandler(args.log_file, encoding='utf-8'),  # 寫入檔案
            logging.StreamHandler()  # 同時顯示在終端
        ]
    )
    
    logging.info("=" * 60)
    logging.info("OpenAI 公告分析器開始執行")
    logging.info(f"執行參數: config={args.config}, limit={args.limit}, output_dir={args.output_dir}")
    logging.info(f"分析類型: {', '.join(args.analysis_types)}")
    logging.info(f"RULC 查詢範圍: {', '.join(map(str, args.rulc))}")
    logging.info(f"測試模式: {'啟用' if args.test_mode else '關閉'}")
    logging.info(f"限速處理機制: 啟用 (最大重試 {4} 次，連續成功 {5} 次重置)")
    logging.info(f"日誌檔案: {args.log_file}")
    logging.info("=" * 60)
    
    try:
        # 初始化分析器（傳入測試模式參數）
        analyzer = OpenAIAnalyzer(config_file=args.config, test_mode=args.test_mode)
        
        if args.test_mode:
            logging.info("啟用測試模式：將記錄所有與OpenAI的對話內容")
        
        # 連接資料庫
        mysql_handler = MySQLHandler(config_file=args.config)
        if not mysql_handler.connect():
            logging.error("連接 MySQL 失敗")
            return
            
        # 選擇資料庫
        if not mysql_handler.select_database():
            logging.error("選擇資料庫失敗")
            return
            
        # 取得指定 RULC 的公告
        announcements = analyzer.get_cl1_announcements(mysql_handler, limit=args.limit, rulc_values=args.rulc)
        if not announcements:
            logging.error(f"未找到 RULC IN ({','.join(map(str, args.rulc))}) 的公告")
            return
            
        logging.info(f"找到 {len(announcements)} 筆 RULC IN ({','.join(map(str, args.rulc))}) 公告，開始分析...")
        
        # 處理公告
        start_time = datetime.now()
        analyzer.process_announcements(announcements, mysql_handler, output_dir=args.output_dir, analysis_types=args.analysis_types)
        end_time = datetime.now()
        
        mysql_handler.close()
        
        processing_time = end_time - start_time
        logging.info("=" * 60)
        logging.info("分析完成")
        logging.info(f"總處理時間: {processing_time}")
        logging.info(f"平均每筆公告處理時間: {processing_time / len(announcements)}")
        logging.info(f"限速統計: 總計 {analyzer.rate_limit_count} 次，連續成功 {analyzer.consecutive_success} 次")
        logging.info(f"日誌檔案已儲存至: {args.log_file}")
        logging.info("=" * 60)
        
    except Exception as e:
        logging.error(f"程式執行失敗: {e}")
        logging.exception("詳細錯誤資訊:")
        raise

if __name__ == '__main__':
    main()
